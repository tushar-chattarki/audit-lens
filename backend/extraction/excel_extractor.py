from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def extract_excel_raw(file_path: str | Path) -> dict[str, Any]:
    """
    Extract raw worksheet data from an Excel workbook.

    This function intentionally does not perform canonical mapping.
    It preserves sheet names, row numbers, column numbers, cell
    references, and raw cell values.
    """

    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(
            f"Unsupported Excel file type: {file_path.suffix}. "
            "Expected .xlsx or .xlsm."
        )

    workbook = load_workbook(
        filename=file_path,
        data_only=True,
    )

    sheets: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        rows: list[dict[str, Any]] = []

        for row in worksheet.iter_rows():
            cells: dict[str, Any] = {}

            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value

            # Preserve only rows containing data.
            if cells:
                rows.append(
                    {
                        "row_index": row[0].row,
                        "cells": cells,
                    }
                )

        sheets.append(
            {
                "sheet_name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "rows": rows,
            }
        )

    return {
        "doc_id": file_path.name,
        "source_type": "xlsx",
        "sheets": sheets,
    }


def save_raw_extraction(
    extraction: dict[str, Any],
    output_path: str | Path,
) -> None:
    """Save raw Excel extraction as JSON."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as file:
        json.dump(
            extraction,
            file,
            indent=2,
            ensure_ascii=False,
            default=str,
        )


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 2:
        print(
            "Usage:\n"
            "python -m backend.extraction.excel_extractor "
            "<excel_file>"
        )
        raise SystemExit(1)

    input_file = Path(sys.argv[1])

    result = extract_excel_raw(input_file)

    output_file = (
        input_file.parent
        / f"{input_file.stem}_raw_extraction.json"
    )

    save_raw_extraction(result, output_file)

    print("Excel extraction successful.")
    print(f"Input: {input_file}")
    print(f"Output: {output_file}")
    print(f"Sheets: {len(result['sheets'])}")

    total_rows = sum(
        len(sheet["rows"])
        for sheet in result["sheets"]
    )

    print(f"Non-empty rows: {total_rows}")

